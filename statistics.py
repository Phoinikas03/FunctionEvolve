import os
import re
import argparse
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from collections import defaultdict
from pathlib import Path
import openpyxl  # 需要先 pip install openpyxl
from openpyxl.styles import Alignment # 引入单元格对齐样式

RANKER_DIR = Path(__file__).resolve().parent / "docs" / "ranker"
if str(RANKER_DIR) not in sys.path:
    sys.path.insert(0, str(RANKER_DIR))

from ranker_experiment import (  # noqa: E402
    apply_complexities,
    compute_complexity_map,
    hit,
    parse_final_candidates,
)
from tune_rankers import (  # noqa: E402
    select_occam_variant,
    select_mdl_variant,
    select_pareto_variant,
)

TOP_KS = (1, 5, 10, 50)
BASELINE_HEURISTIC_EXCLUDED_TASKS = {"direct_prompt", "llmsr", "openevolve"}
BEST_OCCAM5_PARAMS = {
    "log10_delta": 1.0,
    "param_weight": 6.0,
    "special_weight": 2.0,
    "op_weight": 0.25,
    "force_top1": True,
}
BEST_OCCAM10_PARAMS = {
    **BEST_OCCAM5_PARAMS,
}
BEST_PARETO_PARAMS = {
    "param_weight": 6.0,
    "special_weight": 2.0,
    "op_weight": 0.25,
}
BEST_MDL5_PARAMS = {
    "alpha": 0.03,
    "beta": 0.002,
    "gamma": 0.0,
    "op_beta": 0.0,
    "force_top1": False,
}
BEST_MDL10_PARAMS = {
    "alpha": 0.0,
    "beta": 0.0,
    "gamma": 0.0,
    "op_beta": 0.01,
    "force_top1": False,
}
RANKER_METHODS = {
    "occam5": ("occam", {**BEST_OCCAM5_PARAMS, "n": 5}),
    "pareto5": ("pareto", {**BEST_PARETO_PARAMS, "n": 5}),
    "mdl5": ("mdl", {**BEST_MDL5_PARAMS, "n": 5}),
    "occam10": ("occam", {**BEST_OCCAM10_PARAMS, "n": 10}),
    "pareto10": ("pareto", {**BEST_PARETO_PARAMS, "n": 10}),
    "mdl10": ("mdl", {**BEST_MDL10_PARAMS, "n": 10}),
}
DATASET_DIRS = {
    "llm-srbench",
    "llm-srbench-noise1pct",
    "llm-srbench-noise5pct",
    "aifeynman",
}

def natural_sort_key(s):
    """用于自然排序的辅助函数"""
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]

def merge_status(s1, s2):
    """
    合并两个任务的状态。
    优先级规则：正整数(取最小序号) > ? > 0 > 空
    """
    positive_ranks = []
    for status in (s1, s2):
        if is_match_status(status):
            positive_ranks.append(int(status))
    if positive_ranks:
        return str(min(positive_ranks))
    if "?" in (s1, s2):
        return "?"
    if "0" in (s1, s2):
        return "0"
    return ""

def is_match_status(status):
    """正整数表示至少有一个候选匹配，值为最小匹配候选序号。"""
    try:
        return int(status) > 0
    except (TypeError, ValueError):
        return False

def merge_topk_hits(h1, h2):
    """合并 top-k 命中结果：任意一次命中即命中。"""
    merged = {}
    for k in TOP_KS:
        merged[k] = bool((h1 or {}).get(k) or (h2 or {}).get(k))
    return merged

def merge_ranker_hits(h1, h2, methods):
    """合并 ranker 命中结果：任意一次命中即命中。"""
    merged = {}
    for method in methods:
        merged[method] = bool((h1 or {}).get(method) or (h2 or {}).get(method))
    return merged

def parse_matched_candidate_ids(content):
    """解析最后一次验证写入的 Matched candidate IDs。"""
    matched_id_pattern = re.compile(
        r"Matched\s+candidate\s+IDs?\s*[:：]\s*\[([^\]\n]*)\]",
        re.IGNORECASE,
    )
    matched_id_lines = list(matched_id_pattern.finditer(content))
    if not matched_id_lines:
        return None
    return [
        int(value)
        for value in re.findall(r"\d+", matched_id_lines[-1].group(1))
    ]

def parse_candidate_order(content):
    """解析最后一个 Final output 块中候选的原始出现顺序。"""
    lines = content.splitlines()
    start = None
    for i, line in enumerate(lines):
        if re.match(r"\s*Final output\s+\d+\s+formulas", line):
            start = i
    if start is None:
        return []

    candidate_ids = []
    candidate_pattern = re.compile(r"^\s*(\d+)\.\s*(?:\[Mature\]\s*)?<<<")

    for line in lines[start + 1:]:
        if line.startswith("=" * 10):
            break
        candidate_match = candidate_pattern.match(line)
        if candidate_match:
            candidate_ids.append(int(candidate_match.group(1)))

    return candidate_ids

def parse_final_status(content):
    """Parse the final verification result from a log.

    Logs may contain older verification blocks followed by newer ones, so the
    last explicit conclusion/tag is authoritative.
    """
    ids = parse_matched_candidate_ids(content)
    if ids is not None:
        return str(min(ids)) if ids else "0"

    for line in reversed(content.splitlines()):
        if "Conclusion:" in line:
            if "Match found" in line:
                return "1"
            if "No match" in line:
                return "0"
        if "[MATCH FOUND]" in line:
            return "1"
        if "[NO MATCH]" in line:
            return "0"
    return ""

def parse_log_result(content):
    """返回单元格状态和按 log 原始候选顺序计算的 top-k 命中结果。"""
    status = parse_final_status(content)
    matched_ids = parse_matched_candidate_ids(content)
    if matched_ids is None:
        matched_ids = [1] if status == "1" else []

    ranked_ids = parse_candidate_order(content)
    if not ranked_ids:
        ranked_ids = sorted(set(matched_ids))

    matched_set = set(matched_ids)
    topk_hits = {
        k: bool(matched_set.intersection(ranked_ids[:k]))
        for k in TOP_KS
    }
    return status, topk_hits

def should_compute_heuristics(task_name):
    return task_name not in BASELINE_HEURISTIC_EXCLUDED_TASKS

def select_ranker_ids(candidates, selector, params):
    if selector == "occam":
        return select_occam_variant(candidates, **params)
    if selector == "pareto":
        return select_pareto_variant(candidates, **params)
    if selector == "mdl":
        return select_mdl_variant(candidates, **params)
    raise ValueError(selector)

def parse_log_worker(job):
    (
        dataset_name,
        task_name,
        effective_model,
        test_example,
        filepath,
        rel_source,
        include_ranker,
    ) = job
    task_model_key = (effective_model, task_name)
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception as exc:
        return {
            "ok": False,
            "error": f"无法读取文件 {filepath}: {exc}",
        }

    status, topk_hits = parse_log_result(content)
    candidates = parse_final_candidates(content) if include_ranker else []
    matched_ids = None
    if include_ranker and candidates:
        matched_ids = parse_matched_candidate_ids(content)
        if matched_ids is None:
            matched_ids = [1] if status == "1" else []

    return {
        "ok": True,
        "dataset_name": dataset_name,
        "task_model_key": task_model_key,
        "task_name": task_name,
        "effective_model": effective_model,
        "test_example": test_example,
        "status": status,
        "topk_hits": topk_hits,
        "source_file": rel_source,
        "candidates": candidates,
        "matched_ids": matched_ids,
    }

def discover_effective_log_jobs(logs_dir, base_dir):
    """Discover the log file for each dataset/task/model/test case."""
    jobs_by_key = {}
    task_models_by_dataset = defaultdict(set)

    for dataset_name in sorted(os.listdir(logs_dir), key=natural_sort_key):
        if dataset_name not in DATASET_DIRS:
            continue
        dataset_path = os.path.join(logs_dir, dataset_name)
        if not os.path.isdir(dataset_path):
            continue

        for task_name in sorted(os.listdir(dataset_path), key=natural_sort_key):
            task_path = os.path.join(dataset_path, task_name)
            if not os.path.isdir(task_path):
                continue

            for model_name in sorted(os.listdir(task_path), key=natural_sort_key):
                model_path = os.path.join(task_path, model_name)
                if not os.path.isdir(model_path):
                    continue
                if "legacy" in model_name.lower():
                    continue

                effective_model = "no-llm" if task_name.startswith("only_structure") else model_name
                task_model_key = (effective_model, task_name)
                task_models_by_dataset[dataset_name].add(task_model_key)

                for filename in sorted(os.listdir(model_path), key=natural_sort_key):
                    if not filename.endswith((".txt", ".log")):
                        continue
                    filepath = os.path.join(model_path, filename)
                    if not os.path.isfile(filepath):
                        continue
                    filename_without_ext = os.path.splitext(filename)[0]
                    # Run logs may be stored with or without the
                    # _YYYYMMDD_HHMMSS timestamp; strip it when present.
                    # Otherwise drop a baseline method suffix (_direct/_llmsr)
                    # so the equation name is recovered either way.
                    ts_match = re.match(r"^(.+)_\d{8}_\d{6}$", filename_without_ext)
                    if ts_match:
                        test_example = ts_match.group(1)
                    elif filename_without_ext.endswith("_direct"):
                        test_example = filename_without_ext[: -len("_direct")]
                    elif filename_without_ext.endswith("_llmsr"):
                        test_example = filename_without_ext[: -len("_llmsr")]
                    else:
                        test_example = filename_without_ext
                    key = (dataset_name, task_model_key, test_example)
                    jobs_by_key[key] = (
                        dataset_name,
                        task_name,
                        effective_model,
                        test_example,
                        filepath,
                        os.path.relpath(filepath, base_dir),
                        should_compute_heuristics(task_name),
                    )
    return list(jobs_by_key.values()), task_models_by_dataset

def parse_effective_logs(jobs, max_workers):
    if not jobs:
        return []
    worker_count = max(1, max_workers)
    print(f"Parsing {len(jobs)} effective logs with {worker_count} workers", flush=True)
    parsed = []
    if worker_count == 1:
        for i, job in enumerate(jobs, start=1):
            parsed.append(parse_log_worker(job))
            if i % 250 == 0:
                print(f"  parsed logs: {i}/{len(jobs)}", flush=True)
        return parsed

    with ProcessPoolExecutor(max_workers=worker_count) as executor:
        futures = [executor.submit(parse_log_worker, job) for job in jobs]
        for i, future in enumerate(as_completed(futures), start=1):
            parsed.append(future.result())
            if i % 250 == 0:
                print(f"  parsed logs: {i}/{len(jobs)}", flush=True)
    return parsed

def compute_ranker_results(ranker_records, max_workers=1):
    """实时计算 tuned Occam/Pareto/MDL heuristic 命中结果。

    ranker_records maps (test_example, task_model_key) to a dict containing
    content and source metadata. Baseline tasks are filtered before this step.
    """
    parsed = []
    expressions = []
    for key, record in sorted(ranker_records.items(), key=lambda item: (natural_sort_key(item[0][0]), item[0][1])):
        content = record["content"]
        candidates = parse_final_candidates(content)
        if not candidates:
            continue
        expressions.extend(cand.expression for cand in candidates)
        parsed.append((key, content, candidates))

    ranker_results = defaultdict(dict)
    if not parsed:
        return list(RANKER_METHODS), ranker_results

    complexity_map = compute_complexity_map(expressions, max_workers=max_workers)
    for (test_example, task_model_key), content, raw_candidates in parsed:
        candidates = apply_complexities(raw_candidates, complexity_map)
        matched_ids = parse_matched_candidate_ids(content)
        if matched_ids is None:
            matched_ids = [1] if parse_final_status(content) == "1" else []
        matched_set = set(matched_ids)

        hits = {}
        for method, (selector, params) in RANKER_METHODS.items():
            selected_ids = select_ranker_ids(candidates, selector, params)
            hits[method] = hit(selected_ids, matched_set)
        ranker_results[test_example][task_model_key] = hits

    return list(RANKER_METHODS), ranker_results

def compute_ranker_results_from_entries(parsed_entries, complexity_map):
    ranker_results_by_dataset = defaultdict(lambda: defaultdict(dict))
    for entry in parsed_entries:
        candidates = entry.get("candidates") or []
        if not candidates:
            continue
        candidates = apply_complexities(candidates, complexity_map)
        matched_set = set(entry.get("matched_ids") or [])
        hits = {}
        for method, (selector, params) in RANKER_METHODS.items():
            selected_ids = select_ranker_ids(candidates, selector, params)
            hits[method] = hit(selected_ids, matched_set)
        ranker_results_by_dataset[entry["dataset_name"]][entry["test_example"]][entry["task_model_key"]] = hits
    return ranker_results_by_dataset

def write_to_sheet(sheet, test_examples, categories, task_models_list, results, topk_results, ranker_results=None, ranker_methods=None):
    """统一的向 Excel Sheet 写入数据的函数"""
    ranker_results = ranker_results or {}
    ranker_methods = ranker_methods or []
    global_total = len(test_examples)
    global_topk_counts = {
        k: {tm: 0 for tm in task_models_list}
        for k in TOP_KS
    }
    global_ranker_counts = {
        method: {tm: 0 for tm in task_models_list}
        for method in ranker_methods
    }
    global_correct_counts = {tm: 0 for tm in task_models_list}
    for test_example in test_examples:
        for tm in task_models_list:
            topk_hit = topk_results[test_example].get(tm, {})
            for k in TOP_KS:
                if topk_hit.get(k, False):
                    global_topk_counts[k][tm] += 1
            ranker_hit = ranker_results[test_example].get(tm, {})
            for method in ranker_methods:
                if ranker_hit.get(method, False):
                    global_ranker_counts[method][tm] += 1
            if is_match_status(results[test_example].get(tm, "")):
                global_correct_counts[tm] += 1

    # 写入双层表头
    header1 = ['Test Case']
    header2 = ['']
    for model_name, task_name in task_models_list:
        header1.append(model_name)
        header2.append(task_name)

    sheet.append(header1)
    sheet.append(header2)

    # 合并第一行的模型名称单元格 (顶层分栏)
    start_col = 2
    for col in range(3, len(header1) + 2):
        # 当到达列表末尾，或者当前单元格与起始单元格内容不同时，进行合并
        if col > len(header1) or header1[col-1] != header1[start_col-1]:
            if col - 1 > start_col:
                sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col-1)
            start_col = col

    # 合并 A1 和 A2 的 'Test Case' 表头
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # 设置表头居中对齐，看起来更整齐
    center_alignment = Alignment(horizontal='center', vertical='center')
    for r in range(1, 3):
        for c in range(1, len(header1) + 1):
            sheet.cell(row=r, column=c).alignment = center_alignment

    # 写入所有测试集综合的统计行，放在任务名表头下面、具体测试用例前面
    for k in TOP_KS:
        global_summary_row = [f"total_top{k}"]
        for tm in task_models_list:
            global_summary_row.append(f"{global_topk_counts[k][tm]}/{global_total}")
        sheet.append(global_summary_row)

    for method in ranker_methods:
        global_summary_row = [f"total_{method}"]
        for tm in task_models_list:
            global_summary_row.append(f"{global_ranker_counts[method][tm]}/{global_total}")
        sheet.append(global_summary_row)

    global_summary_row = ["total_all"]
    for tm in task_models_list:
        global_summary_row.append(f"{global_correct_counts[tm]}/{global_total}")
    sheet.append(global_summary_row)

    # 遍历排好序的分类
    for cat in sorted(categories.keys(), key=natural_sort_key):
        examples = sorted(categories[cat], key=natural_sort_key)
        cat_total = len(examples)
        
        correct_counts = {tm: 0 for tm in task_models_list}

        for test_example in examples:
            row = [test_example]
            for tm in task_models_list:
                status = results[test_example].get(tm, "")
                row.append(status)
                
                # 统计答对的数量
                if is_match_status(status):
                    correct_counts[tm] += 1
            
            sheet.append(row)
        
        # 写入该分类的总结行
        summary_row = [f"{cat}_all"]
        for tm in task_models_list:
            summary_row.append(f"{correct_counts[tm]}/{cat_total}")
        
        sheet.append(summary_row)

def collect_statistics(max_workers=1, logs_dir=None, base_dir=None):
    """Parse all logs and compute ranker hits across every dataset.

    Returns the in-memory structure ``{"datasets": {...}, "ranker_methods":
    [...]}`` (or ``None`` when there are no logs). This is the parse + ranker
    half of the pipeline, decoupled from xlsx writing so the unified
    orchestrator can consume the results directly instead of round-tripping
    through ``statistics.xlsx`` and re-parsing the logs downstream.

    ``datasets[name]`` holds ``results`` / ``topk_results`` / ``ranker_results``
    keyed by ``test_example`` -> ``(model, task)``, plus ``task_models`` and
    ``test_examples``. ``results[case][(model, task)]`` is exactly the value
    that ``write_to_sheet`` writes into each xlsx cell (the ``combined_status``
    that downstream collection reads back).
    """
    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    if logs_dir is None:
        logs_dir = os.path.join(base_dir, 'logs')
    ranker_methods = list(RANKER_METHODS)

    if not os.path.exists(logs_dir):
        print(f"找不到 logs 目录: {logs_dir}")
        return None

    datasets = {}

    # 1. Discover logs cheaply, then parse effective logs in parallel.
    # Expected layout:
    #   logs/<dataset>/<task>/<model>/<date>/<equation>_<timestamp>.txt
    jobs, task_models_by_dataset = discover_effective_log_jobs(logs_dir, base_dir)
    parsed_entries = parse_effective_logs(jobs, max_workers=max_workers)

    results_by_dataset = defaultdict(lambda: defaultdict(dict))
    topk_by_dataset = defaultdict(lambda: defaultdict(dict))
    test_examples_by_dataset = defaultdict(set)
    ranker_entries = []
    all_ranker_expressions = []

    for entry in parsed_entries:
        if not entry.get("ok"):
            print(entry.get("error", "无法读取日志文件"))
            continue

        dataset_name = entry["dataset_name"]
        test_example = entry["test_example"]
        task_model_key = entry["task_model_key"]
        test_examples_by_dataset[dataset_name].add(test_example)
        results_by_dataset[dataset_name][test_example][task_model_key] = entry["status"]
        topk_by_dataset[dataset_name][test_example][task_model_key] = entry["topk_hits"]

        candidates = entry.get("candidates") or []
        if candidates:
            ranker_entries.append(entry)
            all_ranker_expressions.extend(cand.expression for cand in candidates)

    # 2. Compute expression complexity once globally across all datasets.
    if ranker_entries:
        print(
            f"Computing global complexity for ranker candidates from "
            f"{len(ranker_entries)} logs",
            flush=True,
        )
        complexity_map = compute_complexity_map(
            all_ranker_expressions,
            max_workers=max_workers,
        )
        ranker_results_by_dataset = compute_ranker_results_from_entries(
            ranker_entries,
            complexity_map,
        )
    else:
        ranker_results_by_dataset = defaultdict(lambda: defaultdict(dict))

    for dataset_name in sorted(test_examples_by_dataset.keys(), key=natural_sort_key):
        test_examples = test_examples_by_dataset[dataset_name]
        task_models = task_models_by_dataset.get(dataset_name, set())
        if test_examples and task_models:
            datasets[dataset_name] = {
                "results": results_by_dataset[dataset_name],
                "topk_results": topk_by_dataset[dataset_name],
                "ranker_results": ranker_results_by_dataset[dataset_name],
                "task_models": task_models,
                "test_examples": test_examples,
            }

    if not datasets:
        print("没有在 logs/<dataset> 目录结构下找到有效的日志文件。")
        return None

    return {"datasets": datasets, "ranker_methods": ranker_methods}


def write_statistics_workbook(stats, output_excel):
    """Write the in-memory ``collect_statistics`` structure to an xlsx file."""
    datasets = stats["datasets"]
    ranker_methods = stats["ranker_methods"]

    # 创建 Excel 工作簿；每个数据集一张表。
    wb = openpyxl.Workbook()
    first = True

    def dataset_sheet_sort_key(name):
        return (0, "") if name == "llm-srbench" else (1, natural_sort_key(name))

    for dataset_name in sorted(datasets.keys(), key=dataset_sheet_sort_key):
        data = datasets[dataset_name]
        test_examples = data["test_examples"]
        task_models = data["task_models"]

        categories = defaultdict(list)
        for ex in test_examples:
            match = re.match(r"^([a-zA-Z]+)", ex)
            cat = match.group(1) if match else "Other"
            categories[cat].append(ex)

        task_models_list = sorted(
            list(task_models),
            key=lambda x: (natural_sort_key(x[0]), natural_sort_key(x[1])),
        )

        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = dataset_name[:31]
        write_to_sheet(
            ws,
            test_examples,
            categories,
            task_models_list,
            data["results"],
            data["topk_results"],
            data["ranker_results"],
            ranker_methods,
        )

    # 保存文件
    wb.save(output_excel)


def generate_statistics(max_workers=1):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_excel = os.path.join(base_dir, 'statistics.xlsx')
    stats = collect_statistics(max_workers=max_workers, base_dir=base_dir)
    if stats is None:
        return
    write_statistics_workbook(stats, output_excel)
    print(f"统计完成！数据已保存至 Excel 文件: {output_excel}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--max-workers",
        type=int,
        default=max(1, min(32, os.cpu_count() or 1)),
        help="Parallel workers for tuned Pareto/MDL heuristic complexity computation.",
    )
    args = parser.parse_args()
    generate_statistics(max_workers=args.max_workers)
